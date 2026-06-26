"""
Excel parser for JD upload - Smart multi-sheet handling.

Sheet structure expected:
  Sheet 1 (index 0) : JD criteria — must have skills, good to have, years of exp, etc.
  Sheet 2+ (index 1+): Interview questions captured from past candidates
                        — used as question bank in Interview section

Returns a dict with:
  {
    "jd_text"      : str   — full text of sheet 1 for JD analysis,
    "jd_criteria"  : dict  — parsed key fields from sheet 1,
    "question_bank": list  — questions from sheets 2+ categorized,
    "raw_sheets"   : dict  — raw text per sheet name,
    "all_text"     : str   — combined text of all sheets (fallback),
  }
"""
import os
import re


def _clean(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _is_header(val: str) -> bool:
    """Detect if a cell looks like a column header."""
    val = val.lower()
    keywords = ["question", "skill", "category", "type", "answer",
                "conceptual", "coding", "scenario", "remarks", "level",
                "must", "good", "experience", "years", "requirement"]
    return any(k in val for k in keywords)


def _categorize_question(text: str) -> str:
    """Guess question category from text content."""
    t = text.lower()
    coding_hints    = ["code", "write a", "implement", "algorithm", "function",
                       "program", "debug", "output of", "complexity", "sql",
                       "query", "script", "syntax"]
    scenario_hints  = ["scenario", "situation", "how would you", "what would you",
                       "you are", "imagine", "approach", "handle", "real world",
                       "production", "client", "project", "team", "deadline"]
    if any(h in t for h in coding_hints):
        return "Coding"
    if any(h in t for h in scenario_hints):
        return "Scenario"
    return "Conceptual"


def _parse_xlsx(path: str) -> dict:
    try:
        import openpyxl
    except ImportError:
        return {"error": "INSTALL NEEDED: pip install openpyxl"}

    try:
        wb     = openpyxl.load_workbook(path, data_only=True)
        sheets = wb.worksheets

        if not sheets:
            return {"error": "Excel file has no sheets."}

        # ── SHEET 1: JD criteria ───────────────────────────────────────────
        s1         = sheets[0]
        jd_lines   = [f"=== JD Sheet: {s1.title} ==="]
        jd_criteria = {
            "must_have"  : [],
            "good_to_have": [],
            "experience" : "",
            "role_title" : "",
            "domain"     : "",
            "raw_rows"   : [],
        }

        for row in s1.iter_rows(values_only=True):
            cells = [_clean(c) for c in row if _clean(c)]
            if not cells:
                continue
            row_text = "  |  ".join(cells)
            jd_lines.append(row_text)
            jd_criteria["raw_rows"].append(cells)

            # smart extraction from key-value pairs
            joined = row_text.lower()
            if len(cells) >= 2:
                key = cells[0].lower()
                val = cells[1]

                if any(k in key for k in ["must", "mandatory", "required skill",
                                           "primary", "essential"]):
                    jd_criteria["must_have"].extend(
                        [v.strip() for v in re.split(r"[,\n;]", val) if v.strip()])

                elif any(k in key for k in ["good to have", "nice to have",
                                             "secondary", "preferred", "optional"]):
                    jd_criteria["good_to_have"].extend(
                        [v.strip() for v in re.split(r"[,\n;]", val) if v.strip()])

                elif any(k in key for k in ["experience", "exp", "years"]):
                    jd_criteria["experience"] = val

                elif any(k in key for k in ["role", "position", "title", "designation"]):
                    jd_criteria["role_title"] = val

                elif any(k in key for k in ["domain", "industry", "vertical"]):
                    jd_criteria["domain"] = val

        jd_text = "\n".join(jd_lines)

        # ── SHEETS 2+: Question bank ───────────────────────────────────────
        question_bank   = []
        raw_sheets      = {s1.title: jd_text}

        for sheet in sheets[1:]:
            sheet_lines = [f"=== Question Sheet: {sheet.title} ==="]
            headers     = []
            q_col_idx   = None   # column index that likely holds questions

            for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                cells = [_clean(c) for c in row]

                # detect header row
                if row_idx == 0:
                    headers = cells
                    # find which column has questions
                    for ci, h in enumerate(headers):
                        if h and any(k in h.lower() for k in
                                     ["question", "q ", "interview q", "asked"]):
                            q_col_idx = ci
                            break
                    if q_col_idx is None:
                        # default: first non-empty col with long content is questions
                        q_col_idx = 0
                    sheet_lines.append("  |  ".join([c for c in cells if c]))
                    continue

                non_empty = [c for c in cells if c]
                if not non_empty:
                    continue

                sheet_lines.append("  |  ".join(non_empty))

                # extract question text
                q_text = cells[q_col_idx] if q_col_idx < len(cells) else ""
                if not q_text and non_empty:
                    q_text = non_empty[0]

                if len(q_text) < 8:   # skip very short/empty cells
                    continue

                # detect category from a "category" column or infer from text
                category = "Conceptual"
                cat_col  = None
                for ci, h in enumerate(headers):
                    if h and any(k in h.lower() for k in
                                 ["category", "type", "kind", "section"]):
                        cat_col = ci
                        break
                if cat_col is not None and cat_col < len(cells) and cells[cat_col]:
                    raw_cat  = cells[cat_col].lower()
                    if any(k in raw_cat for k in ["cod", "tech", "program"]):
                        category = "Coding"
                    elif any(k in raw_cat for k in ["scen", "behav", "situation"]):
                        category = "Scenario"
                    else:
                        category = "Conceptual"
                else:
                    category = _categorize_question(q_text)

                # detect difficulty
                difficulty = "Medium"
                for ci, h in enumerate(headers):
                    if h and any(k in h.lower() for k in ["diff", "level", "complex"]):
                        if ci < len(cells) and cells[ci]:
                            dv = cells[ci].lower()
                            if any(k in dv for k in ["easy", "basic", "simple"]):
                                difficulty = "Easy"
                            elif any(k in dv for k in ["hard", "advanced", "expert"]):
                                difficulty = "Hard"
                        break

                question_bank.append({
                    "question"    : q_text,
                    "category"    : category,
                    "difficulty"  : difficulty,
                    "skill_tested": "",
                    "source_sheet": sheet.title,
                    "expected_hints": "",
                })

            raw_sheets[sheet.title] = "\n".join(sheet_lines)

        return {
            "jd_text"      : jd_text,
            "jd_criteria"  : jd_criteria,
            "question_bank": question_bank,
            "raw_sheets"   : raw_sheets,
            "all_text"     : "\n\n".join(raw_sheets.values()),
            "sheet_count"  : len(sheets),
            "q_count"      : len(question_bank),
        }

    except Exception as e:
        return {"error": f"Excel read error (.xlsx): {e}"}


def _parse_xls(path: str) -> dict:
    try:
        import xlrd
    except ImportError:
        return {"error": "INSTALL NEEDED: pip install xlrd"}

    try:
        wb     = xlrd.open_workbook(path)
        sheets = wb.sheets()

        if not sheets:
            return {"error": "Excel file has no sheets."}

        # ── SHEET 1: JD ───────────────────────────────────────────────────
        s1       = sheets[0]
        jd_lines = [f"=== JD Sheet: {s1.name} ==="]
        jd_criteria = {
            "must_have": [], "good_to_have": [],
            "experience": "", "role_title": "", "domain": "", "raw_rows": [],
        }

        for row_idx in range(s1.nrows):
            cells = [_clean(s1.cell_value(row_idx, ci)) for ci in range(s1.ncols)]
            cells = [c for c in cells if c]
            if not cells:
                continue
            row_text = "  |  ".join(cells)
            jd_lines.append(row_text)
            jd_criteria["raw_rows"].append(cells)
            if len(cells) >= 2:
                key = cells[0].lower(); val = cells[1]
                if any(k in key for k in ["must","mandatory","required","primary"]):
                    jd_criteria["must_have"].extend([v.strip() for v in re.split(r"[,\n;]",val) if v.strip()])
                elif any(k in key for k in ["good to have","nice","secondary","preferred"]):
                    jd_criteria["good_to_have"].extend([v.strip() for v in re.split(r"[,\n;]",val) if v.strip()])
                elif any(k in key for k in ["experience","exp","years"]):
                    jd_criteria["experience"] = val
                elif any(k in key for k in ["role","position","title"]):
                    jd_criteria["role_title"] = val
                elif any(k in key for k in ["domain","industry"]):
                    jd_criteria["domain"] = val

        jd_text    = "\n".join(jd_lines)
        raw_sheets = {s1.name: jd_text}
        question_bank = []

        for sheet in sheets[1:]:
            sheet_lines = [f"=== Question Sheet: {sheet.name} ==="]
            headers     = []
            q_col_idx   = 0

            for row_idx in range(sheet.nrows):
                cells = [_clean(sheet.cell_value(row_idx, ci)) for ci in range(sheet.ncols)]
                if row_idx == 0:
                    headers = cells
                    for ci, h in enumerate(headers):
                        if h and "question" in h.lower():
                            q_col_idx = ci; break
                    sheet_lines.append("  |  ".join([c for c in cells if c]))
                    continue
                non_empty = [c for c in cells if c]
                if not non_empty:
                    continue
                sheet_lines.append("  |  ".join(non_empty))
                q_text = cells[q_col_idx] if q_col_idx < len(cells) else ""
                if not q_text and non_empty:
                    q_text = non_empty[0]
                if len(q_text) < 8:
                    continue
                category = _categorize_question(q_text)
                question_bank.append({
                    "question": q_text, "category": category,
                    "difficulty": "Medium", "skill_tested": "",
                    "source_sheet": sheet.name, "expected_hints": "",
                })
            raw_sheets[sheet.name] = "\n".join(sheet_lines)

        return {
            "jd_text": jd_text, "jd_criteria": jd_criteria,
            "question_bank": question_bank, "raw_sheets": raw_sheets,
            "all_text": "\n\n".join(raw_sheets.values()),
            "sheet_count": len(sheets), "q_count": len(question_bank),
        }

    except Exception as e:
        return {"error": f"Excel read error (.xls): {e}"}


def extract_text_from_excel(path: str) -> str:
    """Backward-compatible: returns plain text for JD analysis."""
    result = parse_excel_jd(path)
    if "error" in result:
        return f"[Excel parse error: {result['error']}]"
    return result.get("jd_text", result.get("all_text", ""))


def parse_excel_jd(path: str) -> dict:
    """
    Main function — returns full structured dict.
    Use this in app.py to get both JD text AND question bank.
    """
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        return _parse_xlsx(path)
    elif ext == ".xls":
        return _parse_xls(path)
    else:
        return {"error": f"Unsupported format: {ext}. Use .xlsx or .xls"}
